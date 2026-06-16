from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence


@dataclass
class ImportRow:
    name: str
    hours: float
    cert_no: str = ""


NAME_HEADERS = {
    "学生姓名",
    "姓名",
    "名字",
    "志愿者姓名",
    "志愿者",
    "成员姓名",
    "队员姓名",
    "参与人",
    "服务人员",
    "name",
    "studentname",
}
HOUR_HEADERS = {
    "时数",
    "小时数",
    "认定时数",
    "服务时数",
    "录入时数",
    "志愿时长",
    "服务时长",
    "服务小时",
    "活动时长",
    "工时",
    "hours",
    "hour",
}
CERT_HEADERS = {
    "身份证号",
    "身份证号码",
    "身份证",
    "居民身份证号",
    "居民身份证号码",
    "证件号",
    "证件号码",
    "证件编号",
    "证件",
    "certno",
    "certnumber",
    "idcard",
    "idnumber",
}


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def cell_to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = text.replace("小时", "").replace("时", "").replace("h", "").replace("H", "")
    text = text.replace(",", "").replace("，", "")
    return float(text)


def normalize_header(text: str) -> str:
    return (
        text.strip()
        .lower()
        .replace(" ", "")
        .replace("\u3000", "")
        .replace("_", "")
        .replace("-", "")
        .replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
    )


def find_header(names: list[str], exact: set[str], contains: tuple[str, ...]) -> int | None:
    normalized = [normalize_header(name) for name in names]
    exact_norm = {normalize_header(name) for name in exact}
    for i, name in enumerate(normalized):
        if name in exact_norm:
            return i
    for i, name in enumerate(normalized):
        if any(normalize_header(token) in name for token in contains):
            return i
    return None


def find_columns(header: Sequence[Any]) -> tuple[int, int, int | None]:
    names = [cell_to_text(item) for item in header]
    name_col = find_header(names, NAME_HEADERS, ("姓名", "志愿者", "服务人员", "name"))
    if name_col is None:
        raise ValueError("未找到姓名列：表头建议使用“学生姓名”“姓名”或“志愿者姓名”")

    hour_col = find_header(names, HOUR_HEADERS, ("时数", "小时", "时长", "服务时间", "认定时间", "hours"))
    if hour_col is None:
        raise ValueError("未找到时数列：表头建议使用“认定时数”“服务时数”“志愿时长”或“hours”")

    cert_col = find_header(names, CERT_HEADERS, ("身份证", "证件号", "证件号码", "证件编号", "certno", "idcard", "idnumber"))
    return name_col, hour_col, cert_col


def locate_header(rows: Sequence[Sequence[Any]], max_scan: int = 10) -> tuple[int, int, int, int | None]:
    for index, row in enumerate(rows[:max_scan]):
        try:
            name_col, hour_col, cert_col = find_columns(row)
            return index, name_col, hour_col, cert_col
        except ValueError:
            continue
    preview = []
    for row in rows[: min(len(rows), 3)]:
        preview.append(" | ".join(cell_to_text(cell) for cell in row[:8] if cell_to_text(cell)))
    sample = "\n".join(line for line in preview if line)
    raise ValueError(
        "前 10 行未找到可识别表头。请确认表头包含姓名列和时数列。\n"
        "支持示例：学生姓名 / 姓名 / 志愿者姓名；认定时数 / 服务时数 / 志愿时长。\n"
        + (f"\n当前表格前几行：\n{sample}" if sample else "")
    )


def read_excel(path: str) -> list[ImportRow]:
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(file_path)
    if suffix == ".xlsx":
        return _read_xlsx(file_path)
    raise ValueError("仅支持 .xls 或 .xlsx 文件")


def build_rows(values: Sequence[Sequence[Any]]) -> list[ImportRow]:
    if len(values) < 2:
        return []
    header_index, name_col, hour_col, cert_col = locate_header(values)
    rows: list[ImportRow] = []
    skipped_bad_hours: list[str] = []
    for row in values[header_index + 1 :]:
        row_list = list(row)
        name = cell_to_text(row_list[name_col] if name_col < len(row_list) else "")
        if not name:
            continue
        try:
            hours = cell_to_float(row_list[hour_col] if hour_col < len(row_list) else 0)
        except Exception:
            skipped_bad_hours.append(name)
            continue
        cert_no = cell_to_text(row_list[cert_col] if cert_col is not None and cert_col < len(row_list) else "")
        if hours > 0:
            rows.append(ImportRow(name=name, hours=hours, cert_no=cert_no))
    if not rows and skipped_bad_hours:
        raise ValueError(f"识别到姓名列，但这些行的时数无法转成数字：{', '.join(skipped_bad_hours[:8])}")
    return rows


def _read_xls(path: Path) -> list[ImportRow]:
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("读取 .xls 需要安装 xlrd：pip install xlrd") from exc

    wb = xlrd.open_workbook(str(path))
    sheet = wb.sheet_by_index(0)
    values = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
    return build_rows(values)


def _read_xlsx(path: Path) -> list[ImportRow]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("读取 .xlsx 需要安装 openpyxl：pip install openpyxl") from exc

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[0]
        values = list(sheet.iter_rows(values_only=True))
        return build_rows(values)
    finally:
        wb.close()


def recommended_template_text() -> str:
    return """推荐报表格式

建议直接使用下面这些表头，兼容性最好：

| 学生姓名 | 身份证号 | 认定时数 |
| 张三 | 110101200001010011 | 8 |
| 李四 | 110101200102020022 | 3.5 |

也可以增加备注列，程序会忽略不需要的列。

填写建议：
1. “学生姓名”必须与 bv2008 / 志愿北京实名信息一致。
2. “身份证号”建议保留为文本格式，避免 Excel 自动转成科学计数法或丢失末尾 X。
3. “认定时数”只填数字，例如 8、3.5，不要写“8 小时”。
"""
