"""Batch workflow for recording bv2008 volunteer hours from an XLSX file."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from bv_api import BVApi, PostInfo, RecruitedVolunteer, fetch_all_recruited

ProgressCallback = Callable[[str], None]


@dataclass
class BatchConfig:
    token: str
    activity_id: str
    org_id: str
    start_date: date
    activity_dates: list[str]
    xlsx_path: Path
    output_path: Path
    proof_name: str | None = None
    proof_bytes: bytes | None = None
    proof_mime: str | None = None


REQUIRED_HEADERS = ["姓名", "身份证号", "岗位", "时长", "备注"]
RESULT_HEADER = "录入结果"


def normalize_header(value) -> str:
    return str(value or "").strip()


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_hours(value) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError("时长为空")
    try:
        hours_decimal = Decimal(str(value).strip())
    except InvalidOperation:
        raise ValueError("时长必须是数字") from None
    if not hours_decimal.is_finite():
        raise ValueError("时长必须是数字")
    if hours_decimal <= 0:
        raise ValueError("时长必须大于 0")
    if (hours_decimal * 2) != (hours_decimal * 2).to_integral_value():
        raise ValueError("时长只能填写整数或 .5 小数")
    return float(hours_decimal)


def allocate_hours(hours: float, activity_dates: list[str], max_per_day: float = 10.0) -> list[tuple[str, float]]:
    """Allocate hours across activity dates only.

    Args:
        hours: Total hours to allocate.
        activity_dates: Sorted list of activity date strings (ISO format), e.g. ['2026-06-14', '2026-06-15'].
        max_per_day: Maximum hours per day (default 10).
    """
    if not activity_dates:
        raise ValueError("没有可用的活动日期")
    max_hours = len(activity_dates) * max_per_day
    if hours - max_hours > 1e-9:
        raise ValueError(f"活动共 {len(activity_dates)} 天，最多可录入 {max_hours:g} 小时，不足 {hours:g} 小时")

    out: list[tuple[str, float]] = []
    remaining = hours
    for day_str in activity_dates:
        if remaining <= 1e-9:
            break
        chunk = min(max_per_day, remaining)
        out.append((day_str, chunk))
        remaining -= chunk
    return out


def build_default_notes(times: list[tuple[str, float]], post_name: str) -> str:
    return "，".join(f"{day}{post_name}服务{hour:g}h" for day, hour in times)


def result_output_path(xlsx_path: Path) -> Path:
    return xlsx_path.with_name(f"{xlsx_path.stem}_result{xlsx_path.suffix}")


def build_post_map(posts: list[PostInfo]) -> dict[str, str]:
    post_map: dict[str, str] = {}
    for post in posts:
        post_map[post.name.strip()] = post.post_id
        post_map[post.post_id.strip()] = post.post_id
        if post.post_code:
            post_map[post.post_code.strip()] = post.post_id
    return post_map


def load_sheet(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [normalize_header(cell.value) for cell in ws[1]]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(f"xlsx 缺少列：{', '.join(missing)}")

    index = {header: headers.index(header) + 1 for header in REQUIRED_HEADERS}
    if RESULT_HEADER in headers:
        result_col = headers.index(RESULT_HEADER) + 1
    else:
        result_col = len(headers) + 1
        ws.cell(row=1, column=result_col, value=RESULT_HEADER)
    return wb, ws, index, result_col


def process_row(
    api: BVApi,
    activity_id: str,
    org_id: str,
    post_id: str,
    post_name: str,
    name: str,
    cert_no: str,
    hours: float,
    notes: str,
    activity_dates: list[str],
    proof_name: str | None,
    proof_bytes: bytes | None,
    proof_mime: str | None,
    all_recruited: list[RecruitedVolunteer] | None = None,
    log: Callable[[str], None] | None = None,
) -> str:
    def _log(msg: str) -> None:
        if log:
            log(f"  {msg}")

    times = allocate_hours(hours, activity_dates)
    notes = notes or build_default_notes(times, post_name)

    # Phase 1: query volunteer in org
    user = api.find_org_user(name, cert_no, activity_id, post_id, org_id)
    if user:
        uid = str(user.get("uid", "")).strip()
        if not uid:
            return "失败：查询结果缺少 uid"
        log(f"\n---正在查询志愿者-{name}：已找到（uid={uid}），加入岗位并录入 {hours} 小时")
        api.add_to_post(activity_id, post_id, org_id, uid)
        file_path = api.upload_proof(proof_name, proof_bytes, proof_mime)
        result = api.record_hours(activity_id, post_id, org_id, uid, times, file_path, notes=notes)
        if result.get("resultData") is False:
            return f"失败：录入接口返回失败：{result}"
        return "成功"

    log(f"\n---正在查询志愿者-{name}：未在可招募列表中查到，可能已入岗或未加入团体")

    # Phase 2: not found → try add_member if cert_no available
    if cert_no:
        ok, message = api.add_member(name, cert_no)
        _log(f"  尝试将 {name} 加入团体：{'成功' if ok else '失败'} — {message}")
        if ok:
            user = api.find_org_user(name, cert_no, activity_id, post_id, org_id)
            if user:
                uid = str(user.get("uid", "")).strip()
                if not uid:
                    return "失败：查询结果缺少 uid"
                _log(f"  重新查询志愿者：已找到（uid={uid}），加入岗位并录入 {hours} 小时")
                api.add_to_post(activity_id, post_id, org_id, uid)
                file_path = api.upload_proof(proof_name, proof_bytes, proof_mime)
                result = api.record_hours(activity_id, post_id, org_id, uid, times, file_path, notes=notes)
                if result.get("resultData") is False:
                    return f"失败：录入接口返回失败：{result}"
                return "成功"
            _log("  重新查询志愿者：未查到，用户可能在团体内")

    # Phase 3: user is in the group but not findable via find_org_user
    members = api.find_formal_member(name)
    if len(members) > 1:
        _log(f"  按姓名查找uid：失败，返回 {len(members)} 条结果，请自行确认")
        return f"失败：姓名匹配到 {len(members)} 人，请手动确认后重试"
    if not members:
        _log("  按姓名查找uid：失败，未查到")
        if not cert_no:
            return "失败：按姓名在团体内未查询到志愿者，且缺少身份证号，无法加入团体"
        return "失败：加入团体后仍未查询到 uid"

    member = members[0]
    uid = str(member.get("uid", "")).strip()
    if not uid:
        return "失败：团体成员查询结果缺少 uid"
    _log(f"  按姓名查找uid：成功，找到团体成员（uid={uid}），检查已入岗名单")

    # Check recruited list for 兼项
    if all_recruited:
        for rv in all_recruited:
            if rv.uid == uid:
                if rv.post_id == post_id:
                    _log(f"  已在目标岗位，直接录入 {hours} 小时：录入成功")
                    file_path = api.upload_proof(proof_name, proof_bytes, proof_mime)
                    result = api.record_hours(activity_id, post_id, org_id, uid, times, file_path, notes=notes)
                    if result.get("resultData") is False:
                        return f"失败：录入接口返回失败：{result}"
                    return "成功（已在岗，直接录入）"
                else:
                    _log(f"  已在岗位【{rv.post_name}】，兼项不允许")
                    return f"失败：志愿者已在岗位【{rv.post_name or rv.post_id}】，平台不允许兼项"
        _log("  未在已入岗名单中找到，加入岗位")

    # Not in any recruited list — try to add to post
    api.add_to_post(activity_id, post_id, org_id, uid)
    _log(f"  已加入岗位，录入 {hours} 小时")
    file_path = api.upload_proof(proof_name, proof_bytes, proof_mime)
    result = api.record_hours(activity_id, post_id, org_id, uid, times, file_path, notes=notes)
    if result.get("resultData") is False:
        return f"失败：录入接口返回失败：{result}"
    return "成功"


def run_batch(config: BatchConfig, posts: list[PostInfo], progress: ProgressCallback | None = None) -> Path:
    def log(message: str) -> None:
        if progress:
            progress(message)

    api = BVApi(config.token)
    post_map = build_post_map(posts)
    wb, ws, index, result_col = load_sheet(config.xlsx_path)
    total = max(ws.max_row - 1, 0)
    log(f"读取 {config.xlsx_path.name}，共 {total} 行")

    # Filter activity dates to only those >= start_date
    start_str = config.start_date.isoformat()
    effective_dates = [d for d in config.activity_dates if d >= start_str]
    if not effective_dates:
        raise ValueError(f"没有 >= {start_str} 的活动日期")
    log(f"选择起始日期 {start_str}，可用活动日期 {len(effective_dates)} 天")

    # Pre-fetch all recruited volunteers for fallback matching
    all_recruited: list[RecruitedVolunteer] = []
    try:
        all_recruited = fetch_all_recruited(api, config.activity_id, posts)
        log(f"已获取 {len(all_recruited)} 名已入岗志愿者记录")
        if all_recruited:
            # Group by post name
            by_post: dict[str, list[str]] = {}
            for rv in all_recruited:
                by_post.setdefault(rv.post_name, []).append(rv.name_sensitive)
            for post_name, names in by_post.items():
                log(f"  - {post_name}：{'、'.join(names)}")
    except Exception as exc:
        log(f"获取已入岗名单失败（将继续，但无法检测兼项）：{exc}")

    for row in range(2, ws.max_row + 1):
        name = normalize_text(ws.cell(row=row, column=index["姓名"]).value)
        cert_no = normalize_text(ws.cell(row=row, column=index["身份证号"]).value)
        post_name = normalize_text(ws.cell(row=row, column=index["岗位"]).value)
        notes = normalize_text(ws.cell(row=row, column=index["备注"]).value)
        try:
            hours = parse_hours(ws.cell(row=row, column=index["时长"]).value)
        except Exception as exc:
            result = f"失败：{exc}"
            ws.cell(row=row, column=result_col, value=result)
            log(f"[{row - 1}/{total}] {name or '<空姓名>'}：{result}")
            continue

        if not name or not post_name:
            result = "失败：姓名、岗位均不能为空"
            ws.cell(row=row, column=result_col, value=result)
            log(f"[{row - 1}/{total}] {name or '<空姓名>'}：{result}")
            continue

        post_id = post_map.get(post_name)
        if not post_id:
            result = f"失败：未找到岗位：{post_name}"
            ws.cell(row=row, column=result_col, value=result)
            log(f"[{row - 1}/{total}] {name}：{result}")
            continue

        try:
            result = process_row(
                api=api,
                activity_id=config.activity_id,
                org_id=config.org_id,
                post_id=post_id,
                post_name=post_name,
                name=name,
                cert_no=cert_no,
                hours=hours,
                notes=notes,
                activity_dates=effective_dates,
                proof_name=config.proof_name,
                proof_bytes=config.proof_bytes,
                proof_mime=config.proof_mime,
                all_recruited=all_recruited,
                log=log,
            )
        except Exception as exc:
            result = f"失败：{exc}"

        ws.cell(row=row, column=result_col, value=result)
        log(f"[{row - 1}/{total}] {name} / {post_name} / {hours:g}h：{result}")

    wb.save(config.output_path)
    log(f"结果已写入：{config.output_path}")
    return config.output_path
